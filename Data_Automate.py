import openpyxl

#Openining an excel file or workbook
Scale_Data=openpyxl.load_workbook(r"C:\Users\Kody\Desktop\VALCO\Valco\Scale\Scale project\Scale_Data.xlsx")

#Getting the active sheet
DataSheet=Scale_Data.active

# Iterate over rows in order to most recently filled row
for row in DataSheet.iter_rows():
    # Get the current row index
    row_index = row[0].row

    # Do something with the row index
    print(f"Current row index: {row_index}")

TestData=['1111','20','19:20 PM','12/3/25']
#Rows is used to populate a new row with new data
count=row_index+1

#Populating data
for column in range(1,5,1):
    rows=count
    DataSheet.cell(rows,column).value=TestData[column-1]

#count+=1
#print(count)
print(rows)



Scale_Data.save(r"C:\Users\Kody\Desktop\VALCO\Valco\Scale\Scale project\Scale_Data.xlsx")
Scale_Data.close()



