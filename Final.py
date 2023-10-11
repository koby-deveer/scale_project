import serial
import openpyxl
from datetime import date,datetime
import SQL
import  PrinterConnect

#Setting scale port parameters
ConnectScale= serial.Serial(port='COM6',baudrate=9600,timeout=2)

#Opening Excel workbook
Scale_Data=openpyxl.load_workbook(r"C:\Users\Kody\Desktop\VALCO\Valco\Scale\Scale project\Scale_Data.xlsx")

#Getting the active sheet
DataSheet=Scale_Data.active

# Iterate over rows in order to most recently filled row
for row in DataSheet.iter_rows():
    # Get the current row index
    row_index = row[0].row

    # Do something with the row index
    #print(f"Current row index: {row_index}")

# Since the code above provides the number of the last formatted row, we add one in order to add data to a new row
count=row_index+1

#Check to see if the COM6 port is open, if true then code the while loop will run
while ConnectScale.is_open:

    SerialData=ConnectScale.read(120)
    if len(SerialData)>5:
        InputD=SerialData.decode().splitlines()
        print(InputD)
        dataSize=len(SerialData)
        
        #WeighIn Mode
        if dataSize >1 and dataSize <90:
            #set weigh in mode 
            printerDataIn=SerialData
            Gross=""
            print("Weigh In")
            for item in range(len(InputD)+1):
                match item:
                    case 1:
                        TruckId=InputD[item][11:]

                    case 3:
                        start=InputD[item]

                        for letter in range(len(start)):
                            word=str(start[letter])
                            search=word.isnumeric()

                            if search:
                                Gross+=word

                    case 5:
                        Date=InputD[item][0:7]
                        Time=InputD[item][8:]
            ExData=[TruckId,Gross,Date,Time]

            #For loop to add to the new row and columns
            for column in range(1,5,1):
                rows=count
                DataSheet.cell(rows,column).value=ExData[column-1]
            count+=1

            # Save the data added
            Scale_Data.save(r"C:\Users\Kody\Desktop\VALCO\Valco\Scale\Scale project\Scale_Data.xlsx")

            SetSQL=SQL.SQL_IN(ExData)
            SetPrinterIn=PrinterConnect.Printer(True,printerDataIn)
            print(TruckId)
            print(Gross)
            print(Date)
            print(Time)

        #WeigghOut Mode
        elif dataSize>100:
            #set weighout mode
            printerDataOut=SerialData
            print("Weigh Out")
            Gross=""
            Net=""
            Tare=""
            for item in range(len(InputD)+1):
 
                match item:
                    case 0:
                        TruckId=InputD[item][10:]    
                    
                    case 2:
                        start=InputD[item]

                        for letter in range(len(start)):
                            word=str(start[letter])
                            #print(middle)
                            search=word.isnumeric()
                            if search:
                                Gross+=word

                    case 3:
                        start=InputD[item]

                        for letter in range(len(start)):
                            word=str(start[letter])
                            #print(middle)
                            search=word.isnumeric()
                            if search:
                                Tare+=word
                    
                    case 4:
                        start=InputD[item]

                        for letter in range(len(start)):
                            word=str(start[letter])
                            #print(middle)
                            search=word.isnumeric()
                            if search:
                                Net+=word
                    
                    case 6:
                        Date=InputD[item][0:7]
                        Time=InputD[item][8:]
            ExData=[TruckId,Gross,Tare,Net,Date,Time]

            #For loop to add to the new row and columns
           # for column in range(1,5,1):
            #    rows=count
             #   DataSheet.cell(rows,column).value=ExData[column-1]
            #count+=1

            # Save the data added
            #Scale_Data.save(r"C:\Users\Kody\Desktop\VALCO\Valco\Scale\Scale project\Scale_Data.xlsx")

            SetSQL=SQL.SQL_OUT(ExData)
            SetPrinterOut=PrinterConnect.Printer(True,printerDataOut)

            print(TruckId)
            print(Gross)
            print(Tare)
            print(Net)
            print(Date)
            print(Time)
        

        
    
        # Remove all data in the read buffer
        ConnectScale.reset_input_buffer() 

#Close Scale and Excel file
ConnectScale.close
Scale_Data.close()









